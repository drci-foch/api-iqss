"""
Générateur Excel pour les indicateurs de Lettres de Liaison
Version 1.1 - Avec feuille de graphiques
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from datetime import datetime
from typing import Dict, List, Optional
from io import BytesIO

# --------------------------------------------------------------------
#  CONSTANTES GLOBALES
# --------------------------------------------------------------------

# Couleurs Hôpital Foch (palette officielle)
FOCH_BLUE = "005293"
FOCH_GREEN = "6AA84F"
FOCH_LIGHT_BLUE = "9BC2E6"
FOCH_DARK_BLUE = "003366"
FOCH_GRAY = "595959"

# Couleurs indicateurs
COLOR_GREEN = "92D050"
COLOR_YELLOW = "FFC000"
COLOR_ORANGE = "FF7F27"
COLOR_RED = "FF0000"
COLOR_GRAY = "D9D9D9"
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"


# --------------------------------------------------------------------
#  FONCTIONS UTILITAIRES DE STYLE
# --------------------------------------------------------------------


def get_color_by_threshold(value: float, excellent=95, good=85, medium=70) -> str:
    """Obtenir la couleur selon les seuils"""
    if pd.isna(value):
        return COLOR_GRAY
    if value >= excellent:
        return COLOR_GREEN
    elif value >= good:
        return COLOR_YELLOW
    elif value >= medium:
        return COLOR_ORANGE
    else:
        return COLOR_RED


def apply_cell_style(
    cell,
    font_size=11,
    bold=False,
    font_color=COLOR_BLACK,
    bg_color=None,
    alignment_h="center",
    alignment_v="center",
    border=True,
):
    """Appliquer un style à une cellule"""
    cell.font = Font(name="Calibri", size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(
        horizontal=alignment_h, vertical=alignment_v, wrap_text=True
    )

    if bg_color:
        cell.fill = PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        )

    if border:
        thin_border = Border(
            left=Side(style="thin", color=FOCH_GRAY),
            right=Side(style="thin", color=FOCH_GRAY),
            top=Side(style="thin", color=FOCH_GRAY),
            bottom=Side(style="thin", color=FOCH_GRAY),
        )
        cell.border = thin_border


def set_column_widths(ws, widths):
    """Définir les largeurs de colonnes"""
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width


# --------------------------------------------------------------------
#  FEUILLES EXCEL
# --------------------------------------------------------------------


def create_sheet_resume(
    wb: Workbook, stats_validation: Dict, stats_diffusion: Dict, period: str
):
    """Feuille 1 : Résumé global"""
    ws = wb.create_sheet("Résumé Global", 0)

    # En-tête
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = f"RÉSUMÉ GLOBAL - {period}"
    apply_cell_style(
        cell, font_size=16, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_BLUE
    )

    # Sous-titre
    ws.merge_cells("A2:D2")
    cell = ws["A2"]
    cell.value = "Indicateurs prioritaires : délai de validation des lettres de liaison"
    apply_cell_style(cell, font_size=12, bold=True, bg_color=FOCH_LIGHT_BLUE)

    # Espace
    ws.row_dimensions[3].height = 5

    # En-têtes du tableau
    headers = ["Indicateur", "Valeur", "Objectif", "Statut"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        apply_cell_style(
            cell, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_DARK_BLUE
        )

    # Données
    data_rows = [
        (
            "Nombre total de séjours",
            f"{stats_validation['total_sejours_all']:,}".replace(",", " "),
            "-",
            "📊",
            None,
        ),
        (
            "Taux de validation",
            f"{stats_validation['pct_sejours_validees_all']:.1f}%",
            "≥ 95%",
            "✅"
            if stats_validation["pct_sejours_validees_all"] >= 95
            else "⚠️"
            if stats_validation["pct_sejours_validees_all"] >= 85
            else "❌",
            get_color_by_threshold(
                stats_validation["pct_sejours_validees_all"], 95, 85, 70
            ),
        ),
        (
            "Taux validation J0",
            f"{stats_validation['taux_validation_j0_over_sejours_all']:.1f}%",
            "≥ 90%",
            "✅"
            if stats_validation["taux_validation_j0_over_sejours_all"] >= 90
            else "⚠️"
            if stats_validation["taux_validation_j0_over_sejours_all"] >= 80
            else "❌",
            get_color_by_threshold(
                stats_validation["taux_validation_j0_over_sejours_all"], 90, 80, 70
            ),
        ),
        # (
        #     "Taux diffusion / validation",
        #     f"{stats_diffusion['pct_ll_diffusees_over_validees_all']:.1f}%",
        #     "≥ 90%",
        #     "✅"
        #     if stats_diffusion["pct_ll_diffusees_over_validees_all"] >= 90
        #     else "⚠️"
        #     if stats_diffusion["pct_ll_diffusees_over_validees_all"] >= 80
        #     else "❌",
        #     get_color_by_threshold(
        #         stats_diffusion["pct_ll_diffusees_over_validees_all"], 90, 80, 70
        #     ),
        # ),
    ]

    for row_idx, (indicator, value, objective, status, color) in enumerate(
        data_rows, start=5
    ):
        # Indicateur
        cell = ws.cell(row=row_idx, column=1, value=indicator)
        apply_cell_style(cell, bold=True, alignment_h="left")
        # Valeur
        cell = ws.cell(row=row_idx, column=2, value=value)
        apply_cell_style(cell, bg_color=color if color else None)
        # Objectif
        cell = ws.cell(row=row_idx, column=3, value=objective)
        apply_cell_style(cell)
        # Statut
        cell = ws.cell(row=row_idx, column=4, value=status)
        apply_cell_style(cell, font_size=14)

    # Espace avant la note méthodologique
    ws.row_dimensions[9].height = 10

    # === NOTE MÉTHODOLOGIQUE ===

    # Titre de la note
    ws.merge_cells("A10:D10")
    cell = ws["A10"]
    cell.value = "NOTE MÉTHODOLOGIQUE"
    apply_cell_style(
        cell, font_size=12, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_DARK_BLUE
    )

    # Sous-titre : Typologie des séjours
    ws.merge_cells("A11:D11")
    cell = ws["A11"]
    cell.value = "Typologie des séjours"
    apply_cell_style(cell, font_size=11, bold=True, bg_color=FOCH_LIGHT_BLUE)
    ws.row_dimensions[11].height = 20

    # Contenu méthodologique
    methodology_texts = [
        (
            "• ",
            "Le Décret n° 2016995 du 20 juillet 2016 relatif aux lettres de liaison (NOR : AFSH1612283D) précise que lors de la sortie de l'établissement de santé, une lettre de liaison (LL), rédigée par le médecin de l'établissement qui l'a pris en charge, est remise au patient et transmise le même jour, au médecin traitant.",
        ),
        (
            "• ",
            'Le code de santé publique demande une LL à la sortie de toute "admission" (en opposition aux consultations), HDJ comprises.',
        ),
        ("", ""),
        (
            "📋 ",
            "Séjours pris en compte pour l'indicateur « séjours de 1 nuit et plus » :",
        ),
        ("", "Les séjours suivant sont exclus :"),
        ("      - ", "Patients décédés (séjours non soumis aux LL)"),
        ("      - ", "Chirurgie ambulatoire et Hôpitaux de jours"),
        ("      - ", "Anesthésie, ophtalmologie, radiologie, ORL 392A"),
        ("", ""),
        ("📤 ", "Principe des indicateurs de diffusions (envois) :"),
        (
            "      - ",
            "Seuls les séjours avec lettre de liaison validée par le médecin sont pris en compte",
        ),
        ("      - ", "En excluant :"),
        (
            "            • ",
            "Les LL validées les samedis, dimanche et jours fériés (jours d'absence des secrétaires)",
        ),
        (
            "            • ",
            "Les LL avec plusieurs versions, dont la dernière version est validée à partir de J+1 après la sortie (date de diffusion des versions antérieures non sauvegardées)",
        ),
    ]

    current_row = 12
    for prefix, text in methodology_texts:
        if text == "":  # Ligne vide
            ws.row_dimensions[current_row].height = 5
            current_row += 1
            continue

        ws.merge_cells(f"A{current_row}:D{current_row}")
        cell = ws[f"A{current_row}"]
        cell.value = prefix + text

        # Style différent selon le contenu
        if prefix in ["📋 ", "📤 "]:  # Sous-titres avec émoji
            apply_cell_style(cell, bold=True, alignment_h="left")
            ws.row_dimensions[current_row].height = 30
        elif prefix == "• ":  # Points principaux
            apply_cell_style(cell, alignment_h="left", font_size=10)
            ws.row_dimensions[current_row].height = 40
        else:  # Sous-points
            apply_cell_style(cell, alignment_h="left", font_size=9)
            ws.row_dimensions[current_row].height = 20

        current_row += 1

    # Espace final
    ws.row_dimensions[current_row].height = 10

    # Largeurs de colonnes
    set_column_widths(ws, [35, 15, 15, 10])

    # Ajuster les hauteurs des premières lignes
    for row in range(1, 10):
        ws.row_dimensions[row].height = 25


def create_sheet_validation_detail(
    wb: Workbook, stats_validation: Dict, stats_diffusion: Dict, period: str
):
    """Feuille 2 : Tableau détaillé par spécialité"""
    ws = wb.create_sheet("Détail par Spécialité")

    # En-tête
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = f"Taux de validation et diffusion des LL - SÉJOURS > 24H - {period}"
    apply_cell_style(
        cell, font_size=14, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_BLUE
    )

    # En-têtes du tableau
    headers = [
        "SPÉCIALITÉS",
        "Nb total de séjours",
        "Nb LL validées",
        "% LL validées",
        "Taux de validation à J0 / séjours",
        "Délai validation moyenne)",
        # "Nb LL diffusées",
        # "% des validées",
        # "% des séjours",
        # "Taux de diffusion à J0 de la validation",
        # "Délai diffusions / validation",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_cell_style(
            cell, bold=True, font_color=FOCH_DARK_BLUE, bg_color=FOCH_LIGHT_BLUE
        )

    # Données par spécialité
    specialites_validation = stats_validation.get("par_specialite_all", [])
    # specialites_diffusion = stats_diffusion.get("par_specialite", [])
    # diffusion_dict = {spe["specialite"]: spe for spe in specialites_diffusion}

    for row_idx, spe in enumerate(specialites_validation, start=3):
        # spe_diff = diffusion_dict.get(spe["specialite"], {})

        # Couleur de ligne alternée
        bg_color = COLOR_WHITE if row_idx % 2 == 1 else "F2F2F2"

        # Spécialité
        cell = ws.cell(row=row_idx, column=1, value=spe["specialite"])
        apply_cell_style(
            cell,
            bold=True,
            alignment_h="left",
            bg_color=bg_color,
            font_color=FOCH_DARK_BLUE,
        )

        # Nb total
        cell = ws.cell(row=row_idx, column=2, value=spe["total_sejours"])
        apply_cell_style(cell, bg_color=bg_color)

        # LL valid.
        cell = ws.cell(row=row_idx, column=3, value=spe["nb_sejours_valides"])
        apply_cell_style(cell, bg_color=bg_color)

        # % val.
        pct_val = spe["pct_sejours_validees"]
        cell = ws.cell(row=row_idx, column=4, value=f"{pct_val:.1f}%")
        color_val = get_color_by_threshold(pct_val, 95, 85, 70)
        apply_cell_style(cell, bg_color=color_val)

        # % J0
        pct_j0 = spe["taux_validation_j0_over_sejours"]
        cell = ws.cell(row=row_idx, column=5, value=f"{pct_j0:.1f}%")
        color_j0 = get_color_by_threshold(pct_j0, 90, 80, 70)
        apply_cell_style(cell, bg_color=color_j0)

        # Délai val.
        delai_val = spe.get("delai_moyen_validation", 0)
        if delai_val is None or (isinstance(delai_val, float) and pd.isna(delai_val)):
            delai_val = 0
        cell = ws.cell(row=row_idx, column=6, value=f"{delai_val:.1f}")
        apply_cell_style(cell, bg_color=bg_color)

        # # LL diff.
        # nb_diff = spe_diff.get("nb_ll_diffusees", 0)
        # cell = ws.cell(row=row_idx, column=7, value=nb_diff)
        # apply_cell_style(cell, bg_color=bg_color)

        # # % diff.
        # pct_diff = spe_diff.get("pct_ll_diffusees_over_validees", 0)
        # cell = ws.cell(row=row_idx, column=8, value=f"{pct_diff:.1f}%")
        # color_diff = get_color_by_threshold(pct_diff, 90, 75, 60)
        # apply_cell_style(cell, bg_color=color_diff)

        # # % des séjours
        # pct_diff_sejours = spe_diff.get("pct_ll_diffusees_over_sejours", 0)
        # cell = ws.cell(row=row_idx, column=9, value=f"{pct_diff_sejours:.1f}%")
        # color_diff_global = get_color_by_threshold(pct_diff_sejours, 90, 75, 60)
        # apply_cell_style(cell, bold=True, bg_color=color_diff_global)

        # # Taux de diffusion à J0 de la validation
        # pct_diff_validation = spe_diff.get("taux_diffusion_J0_validation", 0)
        # cell = ws.cell(row=row_idx, column=10, value=f"{pct_diff_validation:.1f}%")
        # color_diff_global = get_color_by_threshold(pct_diff_validation, 90, 75, 60)
        # apply_cell_style(cell, bold=True, bg_color=color_diff_global)

        # # Délai diff. / validation
        # delai_diff_validation = spe_diff.get("delai_diffusion_validation", 0)
        # if delai_diff_validation is None or (
        #     isinstance(delai_diff_validation, float) and pd.isna(delai_diff_validation)
        # ):
        #     delai_diff_validation = 0
        # cell = ws.cell(row=row_idx, column=11, value=f"{delai_diff_validation:.1f}")
        # apply_cell_style(
        #     cell, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_DARK_BLUE
        # )

    # Ligne TOTAL FOCH
    total_row = len(specialites_validation) + 3

    cell = ws.cell(row=total_row, column=1, value="TOTAL FOCH")
    apply_cell_style(
        cell,
        bold=True,
        font_color=COLOR_WHITE,
        bg_color=FOCH_DARK_BLUE,
        alignment_h="left",
    )

    # Nb total de séjours
    cell = ws.cell(
        row=total_row,
        column=2,
        value=f"{stats_validation['total_sejours_all']:,}".replace(",", " "),
    )
    apply_cell_style(cell, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_DARK_BLUE)

    # Nb LL validées
    cell = ws.cell(
        row=total_row,
        column=3,
        value=f"{stats_validation['nb_sejours_valides_all']:,}".replace(",", " "),
    )
    apply_cell_style(cell, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_DARK_BLUE)

    # % LL validées
    pct_global = stats_validation["pct_sejours_validees_all"]
    cell = ws.cell(row=total_row, column=4, value=f"{pct_global:.1f}%")
    color_global = get_color_by_threshold(pct_global, 95, 85, 70)
    apply_cell_style(cell, bold=True, bg_color=color_global)

    # Taux validation à J0 / séjours
    pct_j0_global = stats_validation["taux_validation_j0_over_sejours_all"]
    cell = ws.cell(row=total_row, column=5, value=f"{pct_j0_global:.1f}%")
    color_j0_global = get_color_by_threshold(pct_j0_global, 90, 80, 70)
    apply_cell_style(cell, bold=True, bg_color=color_j0_global)

    # Délai val. moyenne
    delai_global = stats_validation.get("delai_moyen_validation_all", 0)
    if delai_global is None or (
        isinstance(delai_global, float) and pd.isna(delai_global)
    ):
        delai_global = 0
    cell = ws.cell(row=total_row, column=6, value=f"{delai_global:.1f}")
    apply_cell_style(cell, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_DARK_BLUE)

    # # Nb LL diffusées
    # total_diff = stats_diffusion.get("nb_ll_diffusees_all", 0)
    # cell = ws.cell(row=total_row, column=7, value=f"{total_diff:,}".replace(",", " "))
    # apply_cell_style(cell, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_DARK_BLUE)

    # # % des validées
    # pct_diff_global = stats_diffusion.get("pct_ll_diffusees_over_validees_all", 0)
    # cell = ws.cell(row=total_row, column=8, value=f"{pct_diff_global:.1f}%")
    # color_diff_global = get_color_by_threshold(pct_diff_global, 90, 75, 60)
    # apply_cell_style(cell, bold=True, bg_color=color_diff_global)

    # # % des séjours
    # pct_diff_global = stats_diffusion.get("pct_ll_diffusees_over_sejours_all", 0)
    # cell = ws.cell(row=total_row, column=9, value=f"{pct_diff_global:.1f}%")
    # color_diff_global = get_color_by_threshold(pct_diff_global, 90, 75, 60)
    # apply_cell_style(cell, bold=True, bg_color=color_diff_global)

    # # Taux de diffusion à J0 de la validation
    # pct_diff_global = stats_diffusion.get("taux_diffusion_J0_validation_all", 0)
    # cell = ws.cell(row=total_row, column=10, value=f"{pct_diff_global:.1f}%")
    # color_diff_global = get_color_by_threshold(pct_diff_global, 90, 75, 60)
    # apply_cell_style(cell, bold=True, bg_color=color_diff_global)

    # # Délai diff. / validation
    # delai_diff_global = stats_diffusion.get("delai_diffusion_validation_all", 0)
    # if delai_diff_global is None or (
    #     isinstance(delai_diff_global, float) and pd.isna(delai_diff_global)
    # ):
    #     delai_diff_global = 0
    # cell = ws.cell(row=total_row, column=11, value=f"{delai_diff_global:.1f}")
    # apply_cell_style(cell, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_DARK_BLUE)

    # Largeurs de colonnes
    set_column_widths(
        ws,
        [
            25,
            10,
            10,
            10,
            10,
            12,
            #    10,
            #    10,
            #    10,
            #    10,
            #    12
        ],
    )

    # Hauteur des lignes
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 35


def create_sheet_dataframe_analysis(wb: Workbook, df: pd.DataFrame, period: str):
    """Feuille : DataFrame d'analyse brut"""
    ws = wb.create_sheet("Données d'analyse")

    # En-tête
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    cell = ws["A1"]
    cell.value = f"DONNÉES D'ANALYSE - {period}"
    apply_cell_style(
        cell, font_size=14, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_BLUE
    )

    # Sous-titre
    ws.merge_cells(f"A2:{get_column_letter(len(df.columns))}2")
    cell = ws["A2"]
    cell.value = f"Nombre total de lignes : {len(df):,}".replace(",", " ")
    apply_cell_style(cell, font_size=11, bold=True, bg_color=FOCH_LIGHT_BLUE)

    # Espace
    ws.row_dimensions[3].height = 5

    # Convertir le DataFrame en lignes Excel
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=4
    ):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Style pour l'en-tête
            if r_idx == 4:
                apply_cell_style(
                    cell,
                    bold=True,
                    font_color=COLOR_WHITE,
                    bg_color=FOCH_DARK_BLUE,
                    alignment_h="center",
                )
            else:
                # Style alternant pour les données
                bg_color = COLOR_WHITE if r_idx % 2 == 0 else "F8F9FA"
                apply_cell_style(
                    cell,
                    bg_color=bg_color,
                    alignment_h="left" if isinstance(value, str) else "center",
                    font_size=10,
                )

    # Ajuster automatiquement la largeur des colonnes
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = ws.cell(row=4, column=col_idx).column_letter

        for row_idx in range(4, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max(max_length + 2, 12), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Hauteur des lignes d'en-tête
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 30

    # Figer les volets
    ws.freeze_panes = "A5"

    print(
        f"   ↳ Feuille 'Données d'analyse' créée : {len(df)} lignes × {len(df.columns)} colonnes"
    )


# --------------------------------------------------------------------
#  NOUVELLE FEUILLE : GRAPHIQUES
# --------------------------------------------------------------------


def create_sheet_graphiques(
    wb: Workbook,
    stats_validation: Dict,
    stats_diffusion: Dict,
    df_analysis: Optional[pd.DataFrame],
    period: str,
):
    """
    Feuille : Graphiques visuels

    Crée 3 graphiques :
    1. Camembert : Répartition J0 / ≥J1 / Sans LL
    2. Barres horizontales : % LL validées J0 par spécialité
    3. Barres : Nb séjours par spécialité
    """
    ws = wb.create_sheet("Graphiques")

    # ================================================================
    # SECTION 1 : DONNÉES POUR LE CAMEMBERT (Répartition J0/J1+/SansLL)
    # ================================================================

    # Calculer la répartition depuis les stats ou le DataFrame
    if df_analysis is not None and "sej_classe" in df_analysis.columns:
        # Calcul depuis le DataFrame d'analyse
        class_counts = df_analysis["sej_classe"].value_counts()
        nb_j0 = int(class_counts.get("0j", 0))
        nb_j1_plus = int(class_counts.get("1j+", 0))
        nb_sans_ll = int(class_counts.get("sansLL", 0))
    else:
        # Estimation depuis les stats
        total = stats_validation.get("total_sejours_all", 0)
        taux_j0 = stats_validation.get("taux_validation_j0_over_sejours_all", 0)
        taux_val = stats_validation.get("pct_sejours_validees_all", 0)

        nb_j0 = int(total * taux_j0 / 100)
        nb_valides = int(total * taux_val / 100)
        nb_j1_plus = nb_valides - nb_j0
        nb_sans_ll = total - nb_valides

    total_sejours = nb_j0 + nb_j1_plus + nb_sans_ll

    # Calculer les pourcentages
    pct_j0 = (nb_j0 / total_sejours * 100) if total_sejours > 0 else 0
    pct_j1_plus = (nb_j1_plus / total_sejours * 100) if total_sejours > 0 else 0
    pct_sans_ll = (nb_sans_ll / total_sejours * 100) if total_sejours > 0 else 0

    # En-tête de la feuille
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"GRAPHIQUES - INDICATEURS LETTRES DE LIAISON - {period}"
    apply_cell_style(
        cell, font_size=16, bold=True, font_color=COLOR_WHITE, bg_color=FOCH_BLUE
    )
    ws.row_dimensions[1].height = 35

    # ================================================================
    # DONNÉES POUR CAMEMBERT (colonnes A-C, lignes 3-6)
    # ================================================================
    ws["A3"] = "Répartition des séjours"
    apply_cell_style(ws["A3"], bold=True, font_size=12, bg_color=FOCH_LIGHT_BLUE)
    ws.merge_cells("A3:C3")

    # En-têtes
    ws["A4"] = "Catégorie"
    ws["B4"] = "Nombre"
    ws["C4"] = "Pourcentage"
    for col in ["A4", "B4", "C4"]:
        apply_cell_style(
            ws[col], bold=True, bg_color=FOCH_DARK_BLUE, font_color=COLOR_WHITE
        )

    # Données
    pie_data = [
        ("J0", nb_j0, pct_j0),
        ("≥ J1", nb_j1_plus, pct_j1_plus),
        ("Sans LL", nb_sans_ll, pct_sans_ll),
    ]

    for i, (cat, nb, pct) in enumerate(pie_data, start=5):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=nb)
        ws.cell(row=i, column=3, value=round(pct, 1))
        apply_cell_style(ws.cell(row=i, column=1), alignment_h="left")
        apply_cell_style(ws.cell(row=i, column=2))
        apply_cell_style(ws.cell(row=i, column=3))

    # Créer le camembert
    pie_chart = PieChart()
    pie_chart.title = "Répartition des séjours"

    # Références aux données
    labels = Reference(ws, min_col=1, min_row=5, max_row=7)
    data = Reference(ws, min_col=2, min_row=4, max_row=7)

    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)

    # Style du camembert
    pie_chart.width = 12
    pie_chart.height = 8

    # Couleurs personnalisées (vert olive, orange clair, gris)
    colors = ["92D050", "FFC000", "A6A6A6"]
    for i, color in enumerate(colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie_chart.series[0].data_points.append(pt)

    # Étiquettes de données
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showVal = True
    pie_chart.dataLabels.showPercent = True
    pie_chart.dataLabels.showCatName = True

    ws.add_chart(pie_chart, "E3")

    # ================================================================
    # DONNÉES POUR BARRES PAR SPÉCIALITÉ (colonnes A-D, lignes 12+)
    # ================================================================

    # Titre
    ws["A12"] = "Taux de validation J0 par spécialité"
    apply_cell_style(ws["A12"], bold=True, font_size=12, bg_color=FOCH_LIGHT_BLUE)
    ws.merge_cells("A12:D12")

    # En-têtes
    ws["A13"] = "Spécialité"
    ws["B13"] = "Nb séjours"
    ws["C13"] = "% validées"
    ws["D13"] = "% J0"
    for col in ["A13", "B13", "C13", "D13"]:
        apply_cell_style(
            ws[col], bold=True, bg_color=FOCH_DARK_BLUE, font_color=COLOR_WHITE
        )

    # Données par spécialité (triées par taux J0 décroissant pour le graphique)
    specialites = stats_validation.get("par_specialite_all", [])
    specialites_sorted = sorted(
        specialites,
        key=lambda x: x.get("taux_validation_j0_over_sejours", 0),
        reverse=True,
    )

    # Limiter à 15 spécialités pour la lisibilité
    specialites_display = specialites_sorted[:15]

    row_start = 14
    for i, spe in enumerate(specialites_display):
        row = row_start + i
        ws.cell(row=row, column=1, value=spe.get("specialite", ""))
        ws.cell(row=row, column=2, value=spe.get("total_sejours", 0))
        ws.cell(row=row, column=3, value=round(spe.get("pct_sejours_validees", 0), 1))
        ws.cell(
            row=row,
            column=4,
            value=round(spe.get("taux_validation_j0_over_sejours", 0), 1),
        )

        apply_cell_style(ws.cell(row=row, column=1), alignment_h="left")
        apply_cell_style(ws.cell(row=row, column=2))
        apply_cell_style(ws.cell(row=row, column=3))
        apply_cell_style(ws.cell(row=row, column=4))

    row_end = row_start + len(specialites_display) - 1

    # ================================================================
    # GRAPHIQUE BARRES : % Validation J0 par spécialité
    # ================================================================

    bar_chart = BarChart()
    bar_chart.type = "bar"  # Barres horizontales
    bar_chart.style = 10
    bar_chart.title = "% LL validées le jour de la sortie (J0) par service"
    bar_chart.y_axis.title = "Service"
    bar_chart.x_axis.title = "% LL validées J0"

    # Références aux données (colonne D = % J0)
    data_ref = Reference(ws, min_col=4, min_row=13, max_row=row_end)
    cats_ref = Reference(ws, min_col=1, min_row=14, max_row=row_end)

    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    bar_chart.shape = 4

    bar_chart.width = 18
    bar_chart.height = 12

    # Couleur verte pour les barres
    bar_chart.series[0].graphicalProperties.solidFill = "92D050"

    ws.add_chart(bar_chart, "F12")

    # ================================================================
    # GRAPHIQUE BARRES VERTICALES : Nb séjours par spécialité
    # ================================================================

    # Position après le premier graphique de barres
    bar_row = row_end + 5

    ws.cell(row=bar_row, column=1, value="Volume de séjours par spécialité")
    apply_cell_style(
        ws.cell(row=bar_row, column=1),
        bold=True,
        font_size=12,
        bg_color=FOCH_LIGHT_BLUE,
    )
    ws.merge_cells(f"A{bar_row}:D{bar_row}")

    bar_chart2 = BarChart()
    bar_chart2.type = "col"  # Barres verticales
    bar_chart2.style = 10
    bar_chart2.title = "Nombre de séjours par spécialité"
    bar_chart2.x_axis.title = "Spécialité"
    bar_chart2.y_axis.title = "Nombre de séjours"

    # Références aux données (colonne B = Nb séjours)
    data_ref2 = Reference(ws, min_col=2, min_row=13, max_row=row_end)
    cats_ref2 = Reference(ws, min_col=1, min_row=14, max_row=row_end)

    bar_chart2.add_data(data_ref2, titles_from_data=True)
    bar_chart2.set_categories(cats_ref2)

    bar_chart2.width = 18
    bar_chart2.height = 10

    # Couleur bleue Foch pour les barres
    bar_chart2.series[0].graphicalProperties.solidFill = FOCH_BLUE

    ws.add_chart(bar_chart2, f"F{bar_row}")

    # ================================================================
    # AJUSTEMENTS FINAUX
    # ================================================================

    # Largeurs de colonnes
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12

    print(f"   ↳ Feuille 'Graphiques' créée avec 3 graphiques")


# --------------------------------------------------------------------
#  GENERATION DE L'EXCEL
# --------------------------------------------------------------------


def generate_excel(
    stats_validation: Dict,
    stats_diffusion: Dict,
    period: str,
    df_analysis: Optional[pd.DataFrame] = None,
) -> bytes:
    """Générer le fichier Excel avec toutes les feuilles et le retourner en mémoire"""

    wb = Workbook()

    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Créer les feuilles
    create_sheet_resume(wb, stats_validation, stats_diffusion, period)
    create_sheet_validation_detail(wb, stats_validation, stats_diffusion, period)

    # Ajouter la feuille DataFrame si fournie
    if df_analysis is not None and not df_analysis.empty:
        create_sheet_dataframe_analysis(wb, df_analysis, period)

    # NOUVELLE FEUILLE : Graphiques
    create_sheet_graphiques(wb, stats_validation, stats_diffusion, df_analysis, period)

    # Sauvegarder dans un buffer en mémoire
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    print(
        f"✅ Excel généré en mémoire ({len(wb.sheetnames)} feuilles | Formatage harmonisé | Graphiques inclus)"
    )

    return buffer.getvalue()


if __name__ == "__main__":
    # Exemple de test avec données fictives
    test_stats_validation = {
        "total_sejours_all": 1599,
        "nb_sejours_valides_all": 1508,
        "pct_sejours_validees_all": 94.3,
        "taux_validation_j0_over_sejours_all": 81.4,
        "delai_moyen_validation_all": 0.8,
        "par_specialite_all": [
            {
                "specialite": "GERIATRIE",
                "total_sejours": 150,
                "nb_sejours_valides": 150,
                "pct_sejours_validees": 100.0,
                "taux_validation_j0_over_sejours": 100.0,
                "delai_moyen_validation": 0.0,
            },
            {
                "specialite": "THORACIQUE",
                "total_sejours": 120,
                "nb_sejours_valides": 120,
                "pct_sejours_validees": 100.0,
                "taux_validation_j0_over_sejours": 100.0,
                "delai_moyen_validation": 0.0,
            },
            {
                "specialite": "CARDIOLOGIE",
                "total_sejours": 200,
                "nb_sejours_valides": 200,
                "pct_sejours_validees": 100.0,
                "taux_validation_j0_over_sejours": 100.0,
                "delai_moyen_validation": 0.0,
            },
            {
                "specialite": "PNEUMOLOGIE",
                "total_sejours": 90,
                "nb_sejours_valides": 70,
                "pct_sejours_validees": 77.8,
                "taux_validation_j0_over_sejours": 77.8,
                "delai_moyen_validation": 0.5,
            },
            {
                "specialite": "MEDECINE INTERNE",
                "total_sejours": 60,
                "nb_sejours_valides": 40,
                "pct_sejours_validees": 66.7,
                "taux_validation_j0_over_sejours": 66.7,
                "delai_moyen_validation": 0.8,
            },
            {
                "specialite": "ONCOLOGIE",
                "total_sejours": 80,
                "nb_sejours_valides": 40,
                "pct_sejours_validees": 50.0,
                "taux_validation_j0_over_sejours": 50.0,
                "delai_moyen_validation": 1.2,
            },
        ],
    }

    test_stats_diffusion = {
        "nb_ll_diffusees_all": 1508,
        "pct_ll_diffusees_over_validees_all": 100.0,
        "pct_ll_diffusees_over_sejours_all": 94.3,
        "taux_diffusion_J0_validation_all": 85.0,
        "delai_diffusion_validation_all": 0.3,
        "par_specialite": [
            {
                "specialite": "GERIATRIE",
                "nb_ll_diffusees": 150,
                "pct_ll_diffusees_over_validees": 100.0,
                "pct_ll_diffusees_over_sejours": 100.0,
                "taux_diffusion_J0_validation": 95.0,
                "delai_diffusion_validation": 0.1,
            },
            {
                "specialite": "THORACIQUE",
                "nb_ll_diffusees": 120,
                "pct_ll_diffusees_over_validees": 100.0,
                "pct_ll_diffusees_over_sejours": 100.0,
                "taux_diffusion_J0_validation": 90.0,
                "delai_diffusion_validation": 0.2,
            },
            {
                "specialite": "CARDIOLOGIE",
                "nb_ll_diffusees": 200,
                "pct_ll_diffusees_over_validees": 100.0,
                "pct_ll_diffusees_over_sejours": 100.0,
                "taux_diffusion_J0_validation": 92.0,
                "delai_diffusion_validation": 0.2,
            },
            {
                "specialite": "PNEUMOLOGIE",
                "nb_ll_diffusees": 70,
                "pct_ll_diffusees_over_validees": 100.0,
                "pct_ll_diffusees_over_sejours": 77.8,
                "taux_diffusion_J0_validation": 80.0,
                "delai_diffusion_validation": 0.4,
            },
            {
                "specialite": "MEDECINE INTERNE",
                "nb_ll_diffusees": 40,
                "pct_ll_diffusees_over_validees": 100.0,
                "pct_ll_diffusees_over_sejours": 66.7,
                "taux_diffusion_J0_validation": 75.0,
                "delai_diffusion_validation": 0.5,
            },
            {
                "specialite": "ONCOLOGIE",
                "nb_ll_diffusees": 40,
                "pct_ll_diffusees_over_validees": 100.0,
                "pct_ll_diffusees_over_sejours": 50.0,
                "taux_diffusion_J0_validation": 70.0,
                "delai_diffusion_validation": 0.6,
            },
        ],
    }

    # Créer un DataFrame de test
    test_df = pd.DataFrame(
        {
            "sej_id": [f"SEJ{i:04d}" for i in range(100)],
            "sej_classe": ["0j"] * 81 + ["1j+"] * 13 + ["sansLL"] * 6,
            "sej_spe_final": ["CARDIOLOGIE"] * 30
            + ["GERIATRIE"] * 25
            + ["THORACIQUE"] * 20
            + ["PNEUMOLOGIE"] * 15
            + ["ONCOLOGIE"] * 10,
        }
    )

    # Générer le fichier Excel
    excel_bytes = generate_excel(
        test_stats_validation,
        test_stats_diffusion,
        "01/01 au 31/01/2020 (TEST)",
        df_analysis=test_df,
    )

    # Sauvegarder pour test
    with open("test_output_with_charts.xlsx", "wb") as f:
        f.write(excel_bytes)

    print("\n✅ Test terminé ! Fichier sauvegardé : test_output_with_charts.xlsx")
